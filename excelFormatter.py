from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os

def formatter(productName):
	path="/home/runner/JMA-Works/Downloads/"
	path=os.path.join(path,productName+'.xlsx')

	wb2 = load_workbook(path)
	wb2.create_sheet('New Issues')

	sheet = wb2['New Issues']

	thin = Side(border_style="thin", color="000000")
	border = Border(top=thin, left=thin, right=thin, bottom=thin)
	font = Font(bold= True)
	alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
	fill=PatternFill("solid", fgColor="00FFFF00")

	c1 = sheet['A1']
	c1.value = "Sr No."
	c1.alignment = alignment
	c1.font=font
	c1.border=border
	c1.fill=fill

	c1 = sheet['B1']
	c1.value = "Title"
	c1.alignment = alignment
	c1.font=font
	c1.border=border
	c1.fill=fill

	c1 = sheet['C1']
	c1.value = "Path"
	c1.alignment = alignment
	c1.font=font
	c1.border=border
	c1.fill=fill

	c1 = sheet['D1']
	c1.value = "Description"
	c1.alignment = alignment
	c1.font=font
	c1.border=border
	c1.fill=fill
	
	c1 = sheet['E1']
	c1.value = "Device Tested On"
	c1.alignment = alignment
	c1.font=font
	c1.border=border
	c1.fill=fill
	
	c1 = sheet['F1']
	c1.value = "Android Version"
	c1.alignment = alignment
	c1.font=font
	c1.border=border
	c1.fill=fill
	
	c1 = sheet['G1']
	c1.value = "Domain ID"
	c1.alignment = alignment
	c1.font=font
	c1.border=border
	c1.fill=fill
	
	c1 = sheet['H1']
	c1.value = "BOC Validations"
	c1.alignment = alignment
	c1.font=font
	c1.border=border
	c1.fill=fill
	
	c1 = sheet['I1']
	c1.value = "Severity"
	c1.alignment = alignment
	c1.font=font
	c1.border=border
	c1.fill=fill
	
	sheet.column_dimensions['A'].width = 8
	sheet.column_dimensions['B'].width = 35
	sheet.column_dimensions['C'].width = 35
	sheet.column_dimensions['D'].width = 35
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 10
	sheet.column_dimensions['G'].width = 20
	sheet.column_dimensions['H'].width = 20
	sheet.column_dimensions['I'].width = 8
	sheet.row_dimensions[1].height = 30

	c = sheet['J2']
	sheet.freeze_panes = c

	sheet = wb2['Retest Sheet']
	c = sheet['H2']
	sheet.freeze_panes = c

	wb2.save(path)

	