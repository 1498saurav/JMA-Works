from openpyxl import load_workbook
import os

def newIssueSheetCreator(productName):
	path="/home/runner/JMA-Works/Downloads/"
	path=os.path.join(path,productName+'.xlsx')

	wb2 = load_workbook(path)
	wb2.create_sheet('New Issues')

	sheet = wb2['New Issues']

	c1 = sheet['A1']
	c1.value = "Sr No."

	c1 = sheet['B1']
	c1.value = "Title"

	c1 = sheet['C1']
	c1.value = "Path"

	c1 = sheet['D1']
	c1.value = "Description"

	c1 = sheet['E1']
	c1.value = "Device Tested On"

	c1 = sheet['F1']
	c1.value = "Android Version"

	c1 = sheet['G1']
	c1.value = "Domain ID"

	c1 = sheet['H1']
	c1.value = "BOC Validations"

	c1 = sheet['I1']
	c1.value = "Severity"

	wb2.save(path)

	